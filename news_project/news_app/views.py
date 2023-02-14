from django.http import HttpResponse
from openpyxl.workbook import Workbook
from tempfile import NamedTemporaryFile
from rest_framework import viewsets, mixins
from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.viewsets import GenericViewSet

from news_app.models import news_model
from news_app.serializers import NewsSerializer


# Create your views here.
def index_page(request):

    news = news_model.objects.all()

    context = {
                'news' : news,
               }

    return render(request = request, template_name = 'index.html', context = context)


class NewsViewSet( mixins.RetrieveModelMixin,
                   mixins.ListModelMixin,
                   GenericViewSet):

    queryset = news_model.objects.all()
    serializer_class = NewsSerializer


class GenerateExcelView(APIView):

    def get(self, request):
        filename = 'all_news.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.title = "Workbook"

        ws[f'A1'] = 'Заголовок'
        ws[f'B1'] = 'Дата добавления'
        ws[f'C1'] = 'Описание'
        ws[f'D1'] = 'URL-адрес картинки'

        data = news_model.objects.all()
        row_counter = 2
        for line in data:

            ws[f'A{row_counter}'] = line.header
            ws[f'B{row_counter}'] = line.date_created.strftime('%d.%m.%Y')
            ws[f'C{row_counter}'] = line.description
            ws[f'D{row_counter}'] = line.image.url

            row_counter += 1

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content = stream, content_type = 'application/ms-excel')
        # response = HttpResponse(content = stream, content_type = 'application/ms-excel')
        response["Content-Disposition"] = 'attachment; filename="' + filename + '"'
        return response